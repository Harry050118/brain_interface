#!/usr/bin/env python3
"""Generate figures used by the project report from saved experiment outputs."""

from __future__ import annotations

import re
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = ROOT / "eeg_emotion" / "outputs"
FIGURE_DIR = OUTPUT_DIR / "figures"
TRIAL_LOG = OUTPUT_DIR / "logs" / "run_20260414_030257.log"
SUBMISSION_XLSX = OUTPUT_DIR / "submission_bd_conformer_trial_balanced_rank.xlsx"


def parse_subject_accuracies(log_path: Path) -> list[tuple[str, float]]:
    pattern = re.compile(r"\[\d+/60\]\s+(\w+):\s+best_acc=([0-9.]+)")
    rows: list[tuple[str, float]] = []
    for line in log_path.read_text(encoding="utf-8").splitlines():
        match = pattern.search(line)
        if match:
            rows.append((match.group(1), float(match.group(2))))
    if len(rows) != 60:
        raise RuntimeError(f"Expected 60 subject rows in {log_path}, found {len(rows)}")
    return rows


def save_method_comparison() -> None:
    methods = [
        "SVM\nbaseline",
        "Conformer\n+ window norm",
        "Conformer\n+ window rank",
        "Conformer\n+ trial rank",
    ]
    values = np.asarray([65.83, 71.71, 73.80, 84.17])
    colors = ["#7f8c8d", "#4c78a8", "#59a14f", "#e15759"]

    fig, ax = plt.subplots(figsize=(8.0, 4.8), dpi=180)
    bars = ax.bar(methods, values, color=colors, edgecolor="#222222", linewidth=0.8)
    ax.set_ylabel("Full LOSO accuracy (%)")
    ax.set_ylim(50, 90)
    ax.set_title("Performance comparison across model variants")
    ax.grid(axis="y", linestyle="--", linewidth=0.6, alpha=0.45)
    for bar, value in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, value + 0.7, f"{value:.2f}%", ha="center", va="bottom")
    ax.text(
        0.99,
        0.02,
        "Note: trial rank uses trial-level aggregation; the other scores are window-level references.",
        transform=ax.transAxes,
        ha="right",
        va="bottom",
        fontsize=8,
        color="#444444",
    )
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / "method_accuracy_comparison.png", bbox_inches="tight")
    plt.close(fig)


def save_subject_accuracy(rows: list[tuple[str, float]]) -> None:
    subjects = [row[0] for row in rows]
    values = np.asarray([row[1] * 100 for row in rows])
    colors = ["#e15759" if subject.startswith("DEP") else "#4c78a8" for subject in subjects]

    fig, ax = plt.subplots(figsize=(13.5, 5.0), dpi=180)
    ax.bar(np.arange(len(subjects)), values, color=colors, width=0.82)
    ax.axhline(values.mean(), color="#222222", linestyle="--", linewidth=1.2, label=f"Mean = {values.mean():.2f}%")
    ax.set_ylabel("Trial-level LOSO accuracy (%)")
    ax.set_xlabel("Left-out subject")
    ax.set_ylim(0, 105)
    ax.set_xticks(np.arange(len(subjects)))
    ax.set_xticklabels(subjects, rotation=90, fontsize=7)
    ax.set_title("Subject-wise full LOSO performance")
    ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35)
    ax.legend(loc="lower right")
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / "subject_loso_accuracy.png", bbox_inches="tight")
    plt.close(fig)


def save_accuracy_distribution(rows: list[tuple[str, float]]) -> None:
    values = np.asarray([row[1] * 100 for row in rows])

    fig, axes = plt.subplots(1, 2, figsize=(9.5, 4.2), dpi=180)
    axes[0].hist(values, bins=[0, 25, 50, 75, 100, 105], color="#59a14f", edgecolor="#222222")
    axes[0].set_xlabel("Accuracy (%)")
    axes[0].set_ylabel("Number of subjects")
    axes[0].set_title("Accuracy histogram")
    axes[0].grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35)

    axes[1].boxplot(values, vert=True, patch_artist=True, boxprops={"facecolor": "#f28e2b", "alpha": 0.65})
    axes[1].scatter(np.ones_like(values) + np.random.default_rng(42).normal(0, 0.025, size=len(values)), values, s=12, alpha=0.55, color="#333333")
    axes[1].set_xticklabels(["Trial-level\nLOSO"])
    axes[1].set_ylabel("Accuracy (%)")
    axes[1].set_title(f"Mean={values.mean():.2f}%, Std={values.std():.2f}%")
    axes[1].grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35)
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / "accuracy_distribution.png", bbox_inches="tight")
    plt.close(fig)


def save_confusion_matrix(rows: list[tuple[str, float]]) -> None:
    # Each subject has four positive and four neutral trials. Trial-balanced-rank
    # also predicts four positives, so TP and TN are both correct_trials / 2.
    correct_trials = np.asarray([round(row[1] * 8) for row in rows], dtype=int)
    tp = int(correct_trials.sum() // 2)
    tn = int(correct_trials.sum() // 2)
    total_positive = 4 * len(rows)
    total_neutral = 4 * len(rows)
    fn = total_positive - tp
    fp = total_neutral - tn
    matrix = np.asarray([[tn, fp], [fn, tp]], dtype=int)

    fig, ax = plt.subplots(figsize=(4.8, 4.2), dpi=180)
    image = ax.imshow(matrix, cmap="Greens")
    ax.set_xticks([0, 1], labels=["Pred neutral", "Pred positive"])
    ax.set_yticks([0, 1], labels=["True neutral", "True positive"])
    ax.set_title("Trial-level LOSO confusion matrix")
    for i in range(2):
        for j in range(2):
            ax.text(j, i, str(matrix[i, j]), ha="center", va="center", fontsize=15, color="#111111")
    fig.colorbar(image, ax=ax, fraction=0.046, pad=0.04)
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / "trial_level_confusion_matrix.png", bbox_inches="tight")
    plt.close(fig)


def save_submission_distribution(xlsx_path: Path) -> None:
    workbook = load_workbook(xlsx_path, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    subjects = sorted({row[0] for row in rows})
    counts = {subject: {0: 0, 1: 0} for subject in subjects}
    for user_id, _trial_id, label in rows:
        counts[user_id][int(label)] += 1

    neutral = np.asarray([counts[subject][0] for subject in subjects])
    positive = np.asarray([counts[subject][1] for subject in subjects])
    x = np.arange(len(subjects))

    fig, ax = plt.subplots(figsize=(8.5, 4.2), dpi=180)
    ax.bar(x, neutral, label="Pred neutral", color="#4c78a8")
    ax.bar(x, positive, bottom=neutral, label="Pred positive", color="#e15759")
    ax.set_xticks(x)
    ax.set_xticklabels(subjects, rotation=45, ha="right")
    ax.set_ylabel("Number of submitted trials")
    ax.set_ylim(0, 8)
    ax.set_title("Public-test prediction distribution by subject")
    ax.legend(loc="upper right")
    ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35)
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / "submission_prediction_distribution.png", bbox_inches="tight")
    plt.close(fig)


def save_pipeline_diagram() -> None:
    labels = [
        "Raw EEG\n30 x 2500",
        "Window\nnormalization",
        "BD-Conformer\nCNN + attention",
        "Class\nprobabilities",
        "Balanced-rank\npostprocess",
        "Submission\nEmotion_label",
    ]
    x_positions = np.arange(len(labels))
    fig, ax = plt.subplots(figsize=(11.5, 2.7), dpi=180)
    ax.set_axis_off()
    for x, label in zip(x_positions, labels):
        ax.text(
            x,
            0,
            label,
            ha="center",
            va="center",
            fontsize=10,
            bbox={"boxstyle": "round,pad=0.35,rounding_size=0.08", "facecolor": "#f4f6f8", "edgecolor": "#333333"},
        )
    for x in x_positions[:-1]:
        ax.annotate("", xy=(x + 0.72, 0), xytext=(x + 0.28, 0), arrowprops={"arrowstyle": "->", "linewidth": 1.4})
    ax.set_xlim(-0.6, len(labels) - 0.4)
    ax.set_ylim(-0.8, 0.8)
    fig.tight_layout()
    fig.savefig(FIGURE_DIR / "method_pipeline.png", bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    rows = parse_subject_accuracies(TRIAL_LOG)
    save_pipeline_diagram()
    save_method_comparison()
    save_subject_accuracy(rows)
    save_accuracy_distribution(rows)
    save_confusion_matrix(rows)
    save_submission_distribution(SUBMISSION_XLSX)
    print(f"Saved report figures to {FIGURE_DIR}")


if __name__ == "__main__":
    main()
